#!/usr/bin/env python3
"""Extract text (and light structure) from common candidate source files."""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path


def extract_pdf(path: Path) -> str:
    from pypdf import PdfReader

    reader = PdfReader(str(path))
    parts: list[str] = []
    for page in reader.pages:
        t = page.extract_text() or ""
        parts.append(t)
    return "\n\n".join(parts).strip()


def extract_docx(path: Path) -> str:
    import docx

    doc = docx.Document(str(path))
    return "\n".join(p.text for p in doc.paragraphs if p.text.strip()).strip()


def extract_xlsx(path: Path) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(str(path), read_only=True, data_only=True)
    lines: list[str] = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        lines.append(f"## Sheet: {sheet}")
        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip() if c is not None else "" for c in row]
            if any(cells):
                lines.append("\t".join(cells))
        lines.append("")
    wb.close()
    return "\n".join(lines).strip()


def extract_image_meta(path: Path) -> str:
    from PIL import Image

    with Image.open(path) as im:
        fmt = im.format or "unknown"
        mode = im.mode
        size = im.size
    return (
        f"[Image file: {path.name}]\n"
        f"Format: {fmt}, mode: {mode}, size: {size[0]}x{size[1]}\n"
        "Note: OCR is not enabled in this toolchain; describe contents in chat or add a text export."
    )


def extract_text_file(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="replace").strip()


def extract(path: Path) -> tuple[str, str]:
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        return "pdf", extract_pdf(path)
    if suffix == ".docx":
        return "docx", extract_docx(path)
    if suffix in {".xlsx", ".xlsm"}:
        return "xlsx", extract_xlsx(path)
    if suffix in {".png", ".jpg", ".jpeg", ".gif", ".webp", ".tif", ".tiff", ".bmp"}:
        return "image", extract_image_meta(path)
    if suffix == ".doc":
        # Legacy binary Word format: never fall through to the UTF-8 text reader,
        # which would emit mojibake labelled "text" (silent-loss Pitfall 2).
        return "needs-conversion", (
            f"[Legacy binary Word document: {path.name}]\n"
            "Cannot extract text from a .doc file; re-save it as .docx or export to PDF, "
            "then re-run extraction."
        )
    if suffix in {".txt", ".md", ".csv", ".yaml", ".yml", ".json"}:
        return "text", extract_text_file(path)
    try:
        return "text", extract_text_file(path)
    except OSError:
        return "binary", f"[Unreadable as UTF-8 text: {path.name}]"


def main() -> int:
    parser = argparse.ArgumentParser(description="Extract text from PDF, DOCX, XLSX, images, or text files.")
    parser.add_argument("path", type=Path, help="File to extract")
    parser.add_argument(
        "--json",
        action="store_true",
        help="Emit JSON with keys kind, path, content",
    )
    args = parser.parse_args()
    path = args.path.expanduser().resolve()
    if not path.is_file():
        print(f"Not a file: {path}", file=sys.stderr)
        return 1
    kind, content = extract(path)
    if args.json:
        print(json.dumps({"kind": kind, "path": str(path), "content": content}, ensure_ascii=False, indent=2))
    else:
        print(content)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
